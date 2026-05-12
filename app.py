import os
import re

import openpyxl
import requests
from flask import Flask, jsonify, render_template, request

app = Flask(__name__)

# Reused session for all outbound card-image lookups (connection pooling).
_http = requests.Session()
_http.headers["User-Agent"] = "profit-monitor-pkmn/1.0"

XLSM_PATH = os.path.join(os.path.dirname(__file__), "example", "Verdiensten.xlsm")

# ------------------------------------------------------------------
# XLSM helpers
# ------------------------------------------------------------------

def load_transactions():
    wb = openpyxl.load_workbook(XLSM_PATH, keep_vba=True)
    ws = wb["Gegevens"]
    transactions = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 3:
            continue
        type_, amount, description = row[0], row[1], row[2]
        if type_ not in ("Gekocht", "Verkocht") or amount is None:
            continue
        transactions.append(
            {
                "id": row_idx,
                "type": type_,
                "amount": round(float(amount), 2),
                "description": description or "",
            }
        )
    return transactions


def append_transaction(type_: str, amount: float, description: str) -> int:
    wb = openpyxl.load_workbook(XLSM_PATH, keep_vba=True)
    ws = wb["Gegevens"]
    # Enforce sign convention for both types
    amount = abs(amount) if type_ == "Verkocht" else -abs(amount)
    # Find first completely empty row starting from row 3 (rows 1-2 are headers)
    next_row = max(3, ws.max_row + 1)
    for r in range(3, ws.max_row + 2):
        if ws.cell(r, 1).value is None and ws.cell(r, 2).value is None:
            next_row = r
            break
    ws.cell(next_row, 1).value = type_
    ws.cell(next_row, 2).value = round(amount, 2)
    ws.cell(next_row, 3).value = description
    wb.save(XLSM_PATH)
    return next_row


# ------------------------------------------------------------------
# Card name parsing
# ------------------------------------------------------------------

# Pattern matches lines like "Gyarados Lv.52 (STF 19)" or "Turtwig (MEP 040)".
# [^(]+ prevents backtracking over parentheses (avoids ReDoS).
# Single literal space before the parenthesised block avoids polynomial backtracking.
_CARD_PATTERN = re.compile(
    r"^([^(]+) \(([A-Z0-9]+) (\d+[A-Za-z]*)\)$"
)

# Maximum line length fed into the regex to prevent ReDoS on crafted input.
_MAX_CARD_LINE_LEN = 200


def parse_card_lines(description: str) -> list[dict]:
    """
    Parse Pokémon card references from a multi-line description string.

    Each line is expected to follow the format::

        CardName [optional-suffix] (SETCODE NUMBER)

    For example::

        Gyarados Lv.52 (STF 19)
        Turtwig (MEP 040)
        Flareon [4] Lv.55 (RR 60)

    Returns a list of dicts with keys ``name``, ``set``, and ``number``.
    Lines that do not match the pattern are silently skipped.
    """
    cards = []
    for line in description.strip().split("\n"):
        line = line.strip()
        # Truncate pathologically long lines before matching to prevent ReDoS.
        m = _CARD_PATTERN.match(line[:_MAX_CARD_LINE_LEN])
        if not m:
            continue
        raw_name = m.group(1).strip()
        set_code = m.group(2)
        number = m.group(3)
        # Strip optional suffixes from the card name:
        name = re.sub(r" Lv\.\d+$", "", raw_name)      # level indicator, e.g. "Lv.52"
        name = re.sub(r" \[[^\]]{0,20}\]$", "", name)   # bracketed tag, e.g. "[C]" or "[4]"
        # δ Delta Species suffix — use string split to avoid backtracking:
        if " δ" in name:
            name = name[: name.index(" δ")]
        name = name.strip()
        if name:
            cards.append({"name": name, "set": set_code, "number": number})
    return cards


# ------------------------------------------------------------------
# Routes
# ------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/transactions")
def api_transactions():
    return jsonify(load_transactions())


@app.route("/api/summary")
def api_summary():
    txns = load_transactions()
    total_sold = sum(t["amount"] for t in txns if t["type"] == "Verkocht")
    total_bought = sum(t["amount"] for t in txns if t["type"] == "Gekocht")
    profit = total_sold + total_bought
    return jsonify(
        {
            "total_sold": round(total_sold, 2),
            "total_bought": round(total_bought, 2),
            "profit": round(profit, 2),
            "count": len(txns),
        }
    )


@app.route("/api/add", methods=["POST"])
def api_add():
    data = request.get_json(silent=True) or {}
    type_ = data.get("type", "")
    description = (data.get("description") or "").strip()
    try:
        amount = float(data.get("amount", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid amount"}), 400

    if type_ not in ("Gekocht", "Verkocht"):
        return jsonify({"error": "Type must be Gekocht or Verkocht"}), 400
    if amount <= 0:
        return jsonify({"error": "Amount must be a positive number"}), 400

    row = append_transaction(type_, amount, description)
    return jsonify({"success": True, "row": row})


_CM_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
}

_PRODUCT_LINK_RE = re.compile(
    r'href=["\'](?P<url>/en/Pokemon/Products/Singles/[^"\'?#]+)["\']'
)
_PRODUCT_PAGE_RE = re.compile(
    r"/en/Pokemon/Products/Singles/[^/?#]+/[^/?#]+$"
)


def _extract_og_image(html: str) -> str | None:
    """Extract og:image URL regardless of meta attribute order."""
    for pat in (
        r'<meta\s[^>]*property=["\']og:image["\'][^>]*content=["\']([^"\']+)["\']',
        r'<meta\s[^>]*content=["\']([^"\']+)["\'][^>]*property=["\']og:image["\']',
    ):
        m = re.search(pat, html, re.IGNORECASE)
        if m:
            return m.group(1)
    return None


def _cardmarket_image(name: str, set_code: str, number: str) -> str | None:
    """
    Resolve a card image URL via Cardmarket.

    Tries two searches:
    1. Exact name search (may redirect straight to product page).
    2. Name + number search (broader, finds numbered promos).
    For each search result page, follows the first product link and extracts og:image.
    """
    queries = [name, f"{name} {number}"]
    for query in queries:
        url = (
            "https://www.cardmarket.com/en/Pokemon/Products/Singles"
            f"?searchString={requests.utils.quote(query.strip())}"
            "&exactName=false"
        )
        image = _cm_fetch_image(url)
        if image:
            return image
    return None


def _cm_fetch_image(search_url: str) -> str | None:
    """Fetch a Cardmarket search URL and return the first product's og:image."""
    try:
        resp = _http.get(search_url, headers=_CM_HEADERS, timeout=6, allow_redirects=True)
        if not resp.ok:
            return None
        html = resp.text

        # Redirected to a single product page — extract og:image directly.
        if _PRODUCT_PAGE_RE.search(resp.url):
            return _extract_og_image(html)

        # Search results page — follow the first product link.
        m = _PRODUCT_LINK_RE.search(html)
        if not m:
            return None
        prod_resp = _http.get(
            "https://www.cardmarket.com" + m.group("url"),
            headers=_CM_HEADERS,
            timeout=6,
        )
        if prod_resp.ok:
            return _extract_og_image(prod_resp.text)
    except Exception:
        pass
    return None


# ------------------------------------------------------------------
# pokemontcg.io helpers (server-side, avoids browser rate limits)
# ------------------------------------------------------------------

def _tcgq(q: str, page_size: int = 5) -> list[dict]:
    """Query pokemontcg.io and return the data list."""
    try:
        resp = _http.get(
            "https://api.pokemontcg.io/v2/cards",
            params={"q": q, "pageSize": page_size},
            timeout=6,
        )
        if resp.ok:
            return resp.json().get("data", [])
    except Exception:
        pass
    return []


def _lookup_card(name: str, set_code: str, number: str) -> tuple[str | None, str]:
    """
    Full S1-S9 strategy:
    S1-S8 try pokemontcg.io with progressively looser queries.
    S9 falls back to Cardmarket scraping.

    Returns (image_url_or_None, cardmarket_url).
    """
    set_upper = set_code.upper()
    set_lower = set_code.lower()
    set_promo = set_lower + "p"

    # Strip leading zeros: "040" -> "40"
    digits_only = re.sub(r"[^0-9]", "", number) or "0"
    num = str(int(digits_only))
    promo_padded = set_upper + number
    promo_num = set_upper + num

    def first_match(data: list[dict], predicate) -> dict | None:
        return next((c for c in data if predicate(c)), None)

    def num_match(c: dict) -> bool:
        n = c.get("number", "")
        return n in (num, number, promo_num, promo_padded)

    def set_match(c: dict) -> bool:
        s = c.get("set", {})
        ptcgo = s.get("ptcgoCode", "").upper()
        sid = s.get("id", "").upper()
        return ptcgo == set_upper or sid == set_lower.upper() or sid == set_promo.upper()

    hit = None

    # S1: exact — name + number + ptcgoCode
    data = _tcgq(f'name:"{name}" number:{num} set.ptcgoCode:{set_upper}')
    if data:
        hit = data[0]

    # S2: padded promo number + ptcgoCode (e.g. SWSH021)
    if not hit:
        data = _tcgq(f'name:"{name}" number:{promo_padded} set.ptcgoCode:{set_upper}')
        if data:
            hit = data[0]

    # S3: unpadded promo number + ptcgoCode (e.g. SWSH21)
    if not hit and promo_num != promo_padded:
        data = _tcgq(f'name:"{name}" number:{promo_num} set.ptcgoCode:{set_upper}')
        if data:
            hit = data[0]

    # S4: by ptcgoCode, pick number match or first
    if not hit:
        data = _tcgq(f'name:"{name}" set.ptcgoCode:{set_upper}', page_size=20)
        hit = first_match(data, num_match) or (data[0] if data else None)

    # S5: by set.id lowercase
    if not hit:
        data = _tcgq(f'name:"{name}" set.id:{set_lower}', page_size=20)
        hit = first_match(data, num_match) or (data[0] if data else None)

    # S6: by set.id promo variant (e.g. swshp for SWSH promos)
    if not hit:
        data = _tcgq(f'name:"{name}" set.id:{set_promo}', page_size=20)
        hit = first_match(data, num_match) or (data[0] if data else None)

    # S7: wildcard base name + ptcgoCode (handles "Lunala-GX" vs "Lunala GX")
    if not hit:
        base = re.sub(r" (?:GX|EX|V|VMAX|VSTAR)$", "", name).strip()
        if base != name:
            data = _tcgq(f"name:{base}* set.ptcgoCode:{set_upper}", page_size=20)
            hit = first_match(data, num_match)

    # S8: unconstrained, but set code must match
    if not hit:
        data = _tcgq(f'name:"{name}" number:{number}', page_size=20)
        hit = first_match(data, set_match)

    if hit:
        image = hit.get("images", {}).get("small") or None
        # Build a specific Cardmarket search URL using the canonical card name
        cm_url = (
            "https://www.cardmarket.com/en/Pokemon/Products/Search"
            f"?searchString={requests.utils.quote(name)}"
        )
        return image, cm_url

    # S9: Cardmarket scraping fallback
    cm_url = (
        "https://www.cardmarket.com/en/Pokemon/Products/Search"
        f"?searchString={requests.utils.quote(name)}"
    )
    image = _cardmarket_image(name, set_code, number)
    return image, cm_url


@app.route("/api/card-images")
def api_card_images():
    """
    Accepts ?description=... and returns a list of card objects with image URLs.
    Performs full S1-S9 lookup: pokemontcg.io strategies then Cardmarket fallback.
    """
    description = request.args.get("description", "")
    cards = parse_card_lines(description)
    if not cards:
        return jsonify([])

    results = []
    for card in cards:
        image, cm_url = _lookup_card(card["name"], card["set"], card["number"])
        results.append({
            "name": card["name"],
            "set": card["set"],
            "number": card["number"],
            "image": image,
            "cardmarket_url": cm_url,
        })
    return jsonify(results)


if __name__ == "__main__":
    app.run(debug=False)
